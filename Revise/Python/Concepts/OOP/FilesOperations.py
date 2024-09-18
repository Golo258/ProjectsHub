import csv


def writing_to_file(file_path):
    with open(file_path, "w") as file:
        file.write("It's a simple content")


def reading_from_File(file_path):
    with open(file_path, "r") as file:
        for line in file:
            print(line)


def appending_to_File(file_path):
    with open(file_path, "a") as file:
        file.write("\n\nIt's a simple content number 2")


def working_with_csv(file_path):
    data = [
        ['Imię', 'Nazwisko', 'Wiek'],
        ['John', 'Doe', 30],
        ['Jane', 'Doe', 25]
    ]
    with open(file_path, "w", newline='') as csv_file:
        csv_writer = csv.writer(csv_file)
        csv_writer.writerow(data)

    with open(file_path, 'r', newline='') as csv_file2:
        csv_reader = csv.reader(csv_file2)
        for row in csv_reader:
            print(row)


def working_with_json(file_path):
    data = {
        'name': 'John',
        'surname': 'Doe',
        'age': 30
    }
    import json
    with open(file_path, "w") as json_write_file:
        json.dump(data, json_write_file, indent=2)

    with open(file_path, "r") as json_read_file:
        datas = json.load(json_read_file)
        for key, value in datas.items():
            print(key, value)


def working_with_excel(file_path):
    import openpyxl
    #reading
    excel_file = openpyxl.load_workbook(file_path)
    sheet = excel_file.active
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            print("Cell value : " + str(cell.value), end="\t")
    #writing
    new_excel_file  = openpyxl.Workbook()
    new_sheet = new_excel_file.active
    new_sheet['A1'] = 'Imię'
    new_sheet['B1'] = 'Nazwisko'
    new_sheet['A2'] = 'John'
    new_sheet['B2'] = 'Doe'

    new_excel_file.save(file_path)

if __name__ == '__main__':
    file_path = "first.txt"
    csv_file_path = "first.csv"
    json_file_path = "first.json"
    excel_file_path = "first_Excel.xlsx"
    writing_to_file(file_path)
    appending_to_File(file_path)
    reading_from_File(file_path)

    working_with_csv(csv_file_path)
    working_with_json(json_file_path)
    working_with_excel(excel_file_path)
